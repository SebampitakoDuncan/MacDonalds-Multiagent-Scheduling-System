"""
Roster Generator Agent - Exports schedules to Excel format.
"""
import os
from datetime import date, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from .base_agent import BaseAgent
from communication.message import Message, MessageType
from communication.message_bus import MessageBus
from models.schedule import Schedule, Assignment
from models.constraints import ComplianceResult
from models.employee import Employee, EmployeeType, Station
from models.store import Store


class RosterGeneratorAgent(BaseAgent):
    """
    Agent responsible for generating roster output files.
    
    Responsibilities:
    - Export schedule to Excel format
    - Create formatted, printable rosters
    - Generate summary sheets
    - Include compliance reports
    """
    
    def __init__(self, message_bus: MessageBus):
        super().__init__("RosterGenerator", message_bus)
        
        # Style definitions
        self.header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.weekend_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        self.shift_colors = {
            "1F": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # Green
            "2F": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # Yellow
            "3F": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # Orange
            "/": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),   # Gray
        }
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def execute(self,
                schedule: Schedule,
                employees: List[Employee],
                store: Store,
                output_path: str = "output",
                compliance_result: Optional[ComplianceResult] = None,
                **kwargs) -> str:
        """
        Generate Excel roster file.
        
        Args:
            schedule: The schedule to export
            employees: List of employees
            store: Store configuration
            output_path: Directory for output files
            compliance_result: Optional compliance results
            
        Returns:
            Path to the generated file
        """
        self.log(f"Generating roster for {store.name}...")
        
        # Ensure output directory exists
        output_dir = Path(output_path)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Generate filename
        timestamp = date.today().strftime("%Y%m%d")
        filename = f"roster_{store.id}_{schedule.start_date}_{timestamp}.xlsx"
        filepath = output_dir / filename
        
        # Create workbook
        wb = Workbook()
        
        # Main roster sheet
        self._create_roster_sheet(wb, schedule, employees, store)
        
        # Employee summary sheet
        self._create_employee_summary_sheet(wb, schedule, employees)
        
        # Daily coverage sheet
        self._create_coverage_sheet(wb, schedule, store)
        
        # Compliance report sheet
        if compliance_result:
            self._create_compliance_sheet(wb, compliance_result)
        
        # Save workbook
        wb.save(filepath)
        
        # Notify completion
        self.send(
            MessageType.COMPLETE,
            {
                "type": "roster_generated",
                "filepath": str(filepath),
                "sheets": ["Roster", "Employee Summary", "Coverage", "Compliance"],
            },
            receiver="Coordinator"
        )
        
        self.log(f"Roster saved to {filepath}", "success")
        return str(filepath)
    
    def _create_roster_sheet(self, wb: Workbook, schedule: Schedule,
                             employees: List[Employee], store: Store) -> None:
        """Create the main roster sheet."""
        ws = wb.active
        ws.title = "Roster"
        
        # Get date range
        dates = schedule.get_dates_in_range()
        
        # Header row
        headers = ["ID", "Employee Name", "Type", "Station"] + [
            d.strftime("%a\n%d/%m") for d in dates
        ] + ["Total Hours"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        for col in range(5, 5 + len(dates)):
            ws.column_dimensions[chr(64 + col)].width = 8
        
        # Employee rows
        row = 2
        for employee in sorted(employees, key=lambda e: (e.employee_type.value, e.name)):
            ws.cell(row=row, column=1, value=employee.id).border = self.thin_border
            ws.cell(row=row, column=2, value=employee.name).border = self.thin_border
            ws.cell(row=row, column=3, value=employee.employee_type.value).border = self.thin_border
            ws.cell(row=row, column=4, value=employee.primary_station.value).border = self.thin_border
            
            total_hours = 0
            
            for col, target_date in enumerate(dates, 5):
                # Get assignment for this employee on this date
                assignments = [
                    a for a in schedule.get_assignments_by_date(target_date)
                    if a.employee.id == employee.id
                ]
                
                if assignments:
                    assignment = assignments[0]
                    shift_code = assignment.shift.shift_type.value
                    cell = ws.cell(row=row, column=col, value=shift_code)
                    cell.fill = self.shift_colors.get(shift_code, PatternFill())
                    total_hours += assignment.shift.hours
                else:
                    cell = ws.cell(row=row, column=col, value="/")
                    cell.fill = self.shift_colors["/"]
                
                cell.alignment = Alignment(horizontal="center")
                cell.border = self.thin_border
                
                # Weekend highlighting
                if target_date.weekday() >= 5:
                    if not assignments:
                        cell.fill = self.weekend_fill
            
            # Total hours
            hours_cell = ws.cell(row=row, column=5 + len(dates), value=total_hours)
            hours_cell.alignment = Alignment(horizontal="center")
            hours_cell.border = self.thin_border
            
            row += 1
        
        # Legend
        legend_row = row + 2
        ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True)
        legend_items = [
            ("1F", "First Half (06:30-15:30)", "C6EFCE"),
            ("2F", "Second Half (14:00-23:00)", "FFEB9C"),
            ("3F", "Full Day (08:00-20:00)", "FCE4D6"),
            ("/", "Day Off", "D9D9D9"),
        ]
        
        for i, (code, desc, color) in enumerate(legend_items):
            cell = ws.cell(row=legend_row + 1 + i, column=1, value=code)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            ws.cell(row=legend_row + 1 + i, column=2, value=desc)
    
    def _create_employee_summary_sheet(self, wb: Workbook, schedule: Schedule,
                                        employees: List[Employee]) -> None:
        """Create employee summary sheet."""
        ws = wb.create_sheet("Employee Summary")
        
        # Headers
        headers = ["ID", "Name", "Type", "Station", "Shifts", "Week 1 Hours", 
                   "Week 2 Hours", "Total Hours", "Target Min", "Target Max", "Status"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.thin_border
        
        # Data rows
        row = 2
        for employee in sorted(employees, key=lambda e: e.name):
            assignments = schedule.get_assignments_by_employee(employee.id)
            
            # Calculate hours by week
            week1_hours = sum(
                a.shift.hours for a in assignments 
                if (a.shift.date - schedule.start_date).days < 7
            )
            week2_hours = sum(
                a.shift.hours for a in assignments 
                if (a.shift.date - schedule.start_date).days >= 7
            )
            total_hours = week1_hours + week2_hours
            
            # Determine status
            min_target = employee.weekly_hours_target[0] * 2  # 2 weeks
            max_target = employee.weekly_hours_target[1] * 2
            
            if total_hours < min_target:
                status = "⚠️ Under Target"
            elif total_hours > max_target:
                status = "❌ Over Maximum"
            else:
                status = "✓ On Target"
            
            data = [
                employee.id,
                employee.name,
                employee.employee_type.value,
                employee.primary_station.value,
                len(assignments),
                week1_hours,
                week2_hours,
                total_hours,
                employee.weekly_hours_target[0],
                employee.weekly_hours_target[1],
                status
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.thin_border
            
            row += 1
        
        # Auto-size columns
        for col in range(1, 12):
            ws.column_dimensions[chr(64 + col)].width = 14
    
    def _create_coverage_sheet(self, wb: Workbook, schedule: Schedule, 
                                store: Store) -> None:
        """Create daily coverage analysis sheet."""
        ws = wb.create_sheet("Coverage")
        
        # Headers
        headers = ["Date", "Day", "Total Staff", "Kitchen", "Counter"]
        if store.has_mccafe:
            headers.append("McCafe")
        if store.has_dessert_station:
            headers.append("Dessert")
        headers.extend(["Peak Coverage", "Notes"])
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.thin_border
        
        # Data rows
        row = 2
        for target_date in schedule.get_dates_in_range():
            assignments = schedule.get_assignments_by_date(target_date)
            
            # Count by station
            station_counts = {s.value: 0 for s in store.get_active_stations()}
            for a in assignments:
                if a.station.value in station_counts:
                    station_counts[a.station.value] += 1
            
            # Determine peak coverage status
            required = store.get_total_staff_required(is_peak=True)
            if len(assignments) >= required:
                peak_status = "✓ Adequate"
            elif len(assignments) >= required * 0.8:
                peak_status = "⚠️ Marginal"
            else:
                peak_status = "❌ Understaffed"
            
            # Notes
            notes = []
            if target_date.weekday() >= 5:
                notes.append("Weekend")
            if target_date.month == 12 and target_date.day >= 20:
                notes.append("Holiday period")
            
            data = [
                target_date.strftime("%Y-%m-%d"),
                target_date.strftime("%A"),
                len(assignments),
                station_counts.get("Kitchen", 0),
                station_counts.get("Counter", 0),
            ]
            if store.has_mccafe:
                data.append(station_counts.get("Multi-Station McCafe", 0))
            if store.has_dessert_station:
                data.append(station_counts.get("Dessert Station", 0))
            data.extend([peak_status, ", ".join(notes) if notes else ""])
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.thin_border
                
                # Weekend row highlighting
                if target_date.weekday() >= 5:
                    cell.fill = self.weekend_fill
            
            row += 1
        
        # Summary row
        row += 1
        ws.cell(row=row, column=1, value="AVERAGE").font = Font(bold=True)
    
    def _create_compliance_sheet(self, wb: Workbook, 
                                  compliance_result: ComplianceResult) -> None:
        """Create compliance report sheet."""
        ws = wb.create_sheet("Compliance")
        
        # Title
        ws.cell(row=1, column=1, value="COMPLIANCE REPORT").font = Font(bold=True, size=14)
        
        # Summary
        ws.cell(row=3, column=1, value="Overall Status:")
        status = "COMPLIANT ✓" if compliance_result.is_compliant else "NON-COMPLIANT ❌"
        status_cell = ws.cell(row=3, column=2, value=status)
        if compliance_result.is_compliant:
            status_cell.font = Font(color="006400", bold=True)
        else:
            status_cell.font = Font(color="8B0000", bold=True)
        
        ws.cell(row=4, column=1, value="Score:")
        ws.cell(row=4, column=2, value=f"{compliance_result.score:.1f}/100")
        
        ws.cell(row=5, column=1, value="Hard Violations:")
        ws.cell(row=5, column=2, value=len(compliance_result.violations))
        
        ws.cell(row=6, column=1, value="Soft Warnings:")
        ws.cell(row=6, column=2, value=len(compliance_result.warnings))
        
        ws.cell(row=7, column=1, value="Pending Approvals:")
        pending_count = len(compliance_result.pending_approvals)
        ws.cell(row=7, column=2, value=pending_count)
        if pending_count > 0:
            ws.cell(row=7, column=3, value="⚠️ Manager review required")
        
        # Pending Approvals section (Human-in-the-loop)
        if compliance_result.pending_approvals:
            ws.cell(row=9, column=1, value="📋 PENDING MANAGER APPROVAL").font = Font(bold=True, color="0000CD")
            
            headers = ["Type", "Date", "Description", "Reason", "Options"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=10, column=col, value=header).font = Font(bold=True)
            
            row = 11
            for p in compliance_result.pending_approvals:
                ws.cell(row=row, column=1, value=p.constraint_type.value)
                ws.cell(row=row, column=2, value=str(p.affected_date) if p.affected_date else "N/A")
                ws.cell(row=row, column=3, value=p.description)
                ws.cell(row=row, column=4, value=p.details.get("escalation_reason", "")[:60])
                ws.cell(row=row, column=5, value="Accept/Overtime/Casual/Reduce")
                row += 1
            
            row += 1
        else:
            row = 9
        
        # Violations list
        if compliance_result.violations:
            ws.cell(row=row, column=1, value="VIOLATIONS").font = Font(bold=True, color="8B0000")
            row += 1
            
            headers = ["Type", "Severity", "Description", "Affected", "Suggestion"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header).font = Font(bold=True)
            row += 1
            
            for v in compliance_result.violations:
                ws.cell(row=row, column=1, value=v.constraint_type.value)
                ws.cell(row=row, column=2, value=v.severity)
                ws.cell(row=row, column=3, value=v.description)
                ws.cell(row=row, column=4, value=v.affected_entity)
                ws.cell(row=row, column=5, value=v.suggestions[0] if v.suggestions else "")
                row += 1
        
        # Warnings list
        if compliance_result.warnings:
            row += 2
            ws.cell(row=row, column=1, value="WARNINGS").font = Font(bold=True, color="B8860B")
            row += 1
            
            headers = ["Type", "Severity", "Description"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header).font = Font(bold=True)
            row += 1
            
            for w in compliance_result.warnings[:10]:  # Limit to 10
                ws.cell(row=row, column=1, value=w.constraint_type.value)
                ws.cell(row=row, column=2, value=w.severity)
                ws.cell(row=row, column=3, value=w.description)
                row += 1
        
        # Auto-size columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 40
    
    def _on_request(self, message: Message) -> None:
        """Handle requests from other agents."""
        content = message.content
        
        if isinstance(content, dict):
            if content.get("type") == "generate_roster":
                # Extract parameters and generate
                schedule = content.get("schedule")
                employees = content.get("employees")
                store = content.get("store")
                
                if all([schedule, employees, store]):
                    filepath = self.execute(schedule, employees, store)
                    self.respond(message, {"filepath": filepath})

